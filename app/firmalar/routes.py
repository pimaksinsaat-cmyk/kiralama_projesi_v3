import os
from io import BytesIO
from flask import render_template, url_for, redirect, flash, request, current_app, send_file
from datetime import date
from flask_login import current_user, login_required
from decimal import Decimal
from sqlalchemy import asc, desc
from openpyxl import Workbook, load_workbook

from app.firmalar import firmalar_bp
from app.firmalar.forms import FirmaForm
from app.firmalar.models import Firma
from app.extensions import db
from app.services.firma_services import FirmaService
from app.services.base import ValidationError

# --- GÜVENLİK YARDIMCISI ---
def get_actor_id():
    """Kullanıcı giriş yapmışsa ID'sini döner, aksi halde None."""
    return getattr(current_user, 'id', None)


class ListPagination:
    """In-memory list için sayfalama yardımcısı."""
    def __init__(self, total, page, per_page):
        self.total = max(0, int(total or 0))
        self.per_page = max(1, int(per_page or 1))
        self.pages = (self.total + self.per_page - 1) // self.per_page if self.total else 0
        self.page = max(1, min(int(page or 1), self.pages or 1))

    @property
    def has_prev(self): return self.page > 1
    @property
    def has_next(self): return self.page < self.pages
    @property
    def prev_num(self): return self.page - 1
    @property
    def next_num(self): return self.page + 1

    def iter_pages(self, left_edge=1, left_current=1, right_current=2, right_edge=1):
        last = 0
        for num in range(1, self.pages + 1):
            if (
                num <= left_edge
                or (self.page - left_current - 1 < num < self.page + right_current)
                or num > self.pages - right_edge
            ):
                if last + 1 != num:
                    yield None
                yield num
                last = num


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {'1', 'true', 'evet', 'yes', 'x'}:
        return True
    if text in {'0', 'false', 'hayir', 'hayır', 'no', ''}:
        return False
    return default

# -------------------------------------------------------------------------
# 1. Aktif Firma Listeleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/')
@firmalar_bp.route('/index')
@login_required
def index():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        if per_page not in {10, 25, 50, 100}:
            per_page = 50
        q = request.args.get('q', '', type=str)
        sort_by = request.args.get('sort_by', 'firma_adi', type=str)
        sort_dir = request.args.get('sort_dir', 'asc', type=str)

        allowed_sort_fields = {
            'firma_adi': Firma.firma_adi,
            'yetkili_adi': Firma.yetkili_adi,
            'vergi_no': Firma.vergi_no
        }
        if sort_by not in allowed_sort_fields:
            sort_by = 'firma_adi'
        if sort_dir not in ('asc', 'desc'):
            sort_dir = 'asc'
        
        query = FirmaService.get_active_firms(search_query=q)
        sort_column = allowed_sort_fields[sort_by]
        sort_expression = asc(sort_column) if sort_dir == 'asc' else desc(sort_column)
        query = query.order_by(None).order_by(sort_expression, desc(Firma.id))
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return render_template(
            'firmalar/index.html',
            firmalar=pagination.items,
            pagination=pagination,
            per_page=per_page,
            q=q,
            sort_by=sort_by,
            sort_dir=sort_dir
        )
    except Exception as e:
        current_app.logger.error(f"Firma listesi yüklenirken hata: {str(e)}")
        flash("Firma listesi şu an yüklenemiyor.", "danger")
        return redirect(url_for('main.index')) # Ana sayfaya yönlendir

# -------------------------------------------------------------------------
# 2. Pasif (Arşivlenmiş) Firma Listeleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/pasif')
@login_required
def pasif_index():
    try:
        page = request.args.get('page', 1, type=int)
        q = request.args.get('q', '', type=str)
        
        query = FirmaService.get_inactive_firms(search_query=q)
        pagination = query.paginate(page=page, per_page=50, error_out=False)
        
        return render_template('firmalar/pasif_index.html', firmalar=pagination.items, pagination=pagination, q=q)
    except Exception as e:
        current_app.logger.error(f"Pasif firma listesi hatası: {str(e)}")
        flash("Arşivlenmiş firmalar yüklenemedi.", "danger")
        return redirect(url_for('firmalar.index'))

# -------------------------------------------------------------------------
# 3. Yeni Firma Ekleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
def ekle():
    form = FirmaForm()
    
    if form.validate_on_submit():
        try:
            # Servis katmanında tanımlı güncellenebilir alanları süzerek al
            data = {k: v for k, v in form.data.items() if k in FirmaService.updatable_fields}
            
            # Başlangıç değerlerini set et
            data.update({
                'bakiye': Decimal('0'),
                'sozlesme_rev_no': 0,
                'sozlesme_no': None
            })
            
            yeni_firma = Firma(**data)
            FirmaService.save(yeni_firma, actor_id=get_actor_id())
            
            flash(f"'{yeni_firma.firma_adi}' başarıyla sisteme kaydedildi.", "success")
            return redirect(url_for('firmalar.index'))
        except ValidationError as e:
            flash(str(e), "warning")
        except Exception as e:
            current_app.logger.error(f"Firma ekleme hatası: {str(e)}")
            flash("Kayıt sırasında sistemsel bir hata oluştu.", "danger")
            
    return render_template('firmalar/ekle.html', form=form, today_date=date.today().strftime('%d.%m.%Y'))

# -------------------------------------------------------------------------
# 4. Sözleşme Hazırla
# -------------------------------------------------------------------------
@firmalar_bp.route('/sozlesme-hazirla/<int:id>', methods=['POST'])
@login_required
def sozlesme_hazirla(id):
    try:
        app_path = os.path.join(os.getcwd(), 'app')
        FirmaService.sozlesme_hazirla(firma_id=id, base_app_path=app_path, actor_id=get_actor_id())
        flash("Sözleşme numarası başarıyla atandı ve arşiv klasörleri oluşturuldu.", "success")
    except ValidationError as e:
        flash(str(e), "warning")
    except Exception as e:
        current_app.logger.error(f"Sözleşme hazırlama hatası (Firma ID: {id}): {str(e)}")
        flash("Klasör yapısı oluşturulurken bir hata meydana geldi.", "danger")
    return redirect(url_for('firmalar.index'))

# -------------------------------------------------------------------------
# 5. Firma Düzenleme
# -------------------------------------------------------------------------
@firmalar_bp.route('/duzelt/<int:id>', methods=['GET', 'POST'])
@login_required
def duzelt(id):
    firma = FirmaService.get_by_id(id)
    if not firma:
        flash("Düzenlenmek istenen firma bulunamadı!", "danger")
        return redirect(url_for('firmalar.index'))
        
    form = FirmaForm(obj=firma)
    
    # Özel alanları form verisine elle eşle (Form yapısına göre)
    if request.method == 'GET':
        form.genel_sozlesme_no.data = firma.sozlesme_no
        form.sozlesme_rev_no.data = firma.sozlesme_rev_no
        form.sozlesme_tarihi.data = firma.sozlesme_tarihi

    if form.validate_on_submit():
        try:
            data = {k: v for k, v in form.data.items() if k in FirmaService.updatable_fields}
            
            # Formdaki özel alanları veritabanı alanlarıyla eşleştir
            data.update({
                'sozlesme_no': form.genel_sozlesme_no.data,
                'sozlesme_rev_no': form.sozlesme_rev_no.data,
                'sozlesme_tarihi': form.sozlesme_tarihi.data
            })
            
            FirmaService.update(id, data, actor_id=get_actor_id())
            flash(f"'{firma.firma_adi}' bilgileri başarıyla güncellendi.", 'success')
            return redirect(url_for('firmalar.index'))
        except ValidationError as e:
            flash(str(e), "danger")
        except Exception as e:
            current_app.logger.error(f"Firma güncelleme hatası (ID: {id}): {str(e)}")
            flash("Bilgiler güncellenirken bir hata oluştu.", "danger")
            
    return render_template('firmalar/duzelt.html', form=form, firma=firma)

# -------------------------------------------------------------------------
# 6. Firma Bilgi / Cari Ekstre
# -------------------------------------------------------------------------
@firmalar_bp.route('/bilgi/<int:id>', methods=['GET'])
@login_required
def bilgi(id):
    try:
        tab = request.args.get('tab', 'hareket')

        hareket_per_page = request.args.get('hareket_per_page', 20, type=int)
        hareket_page     = request.args.get('hareket_page', 1, type=int)
        kiralama_per_page = request.args.get('kiralama_per_page', 10, type=int)
        kiralama_page     = request.args.get('kiralama_page', 1, type=int)

        allowed_pp = {10, 20, 25, 50, 100}
        if hareket_per_page not in allowed_pp:   hareket_per_page = 20
        if kiralama_per_page not in allowed_pp:  kiralama_per_page = 10

        finans_verileri = FirmaService.get_financial_summary(id)

        # --- Cari Hareketler sayfalama ---
        hareketler_all = finans_verileri.pop('hareketler')
        h_pag = ListPagination(total=len(hareketler_all), page=hareket_page, per_page=hareket_per_page)
        hb = (h_pag.page - 1) * h_pag.per_page
        hareketler_sayfa = hareketler_all[hb: hb + h_pag.per_page]

        # --- Kiralamalar sayfalama ---
        firma = finans_verileri['firma']
        kiralamalar_all = sorted(firma.kiralamalar, key=lambda k: k.id, reverse=True)
        k_pag = ListPagination(total=len(kiralamalar_all), page=kiralama_page, per_page=kiralama_per_page)
        kb = (k_pag.page - 1) * k_pag.per_page
        kiralamalar_sayfa = kiralamalar_all[kb: kb + k_pag.per_page]

        return render_template(
            'firmalar/bilgi.html',
            **finans_verileri,
            hareketler=hareketler_sayfa,
            hareket_pagination=h_pag,
            hareket_per_page=hareket_per_page,
            kiralamalar_sayfa=kiralamalar_sayfa,
            kiralama_pagination=k_pag,
            kiralama_per_page=kiralama_per_page,
            tab=tab,
        )
    except ValidationError as e:
        flash(str(e), "danger")
        return redirect(url_for('firmalar.index'))
    except Exception as e:
        current_app.logger.error(f"Cari özet yükleme hatası (Firma ID: {id}): {str(e)}")
        flash("Finansal veriler şu an hesaplanamıyor.", "danger")
        return redirect(url_for('firmalar.index'))

# -------------------------------------------------------------------------
# 7. Firmayı Arşive Kaldır (Silme - Kontrollü)
# -------------------------------------------------------------------------
@firmalar_bp.route('/sil/<int:id>', methods=['POST'])
@login_required
def sil(id):
    try:
        FirmaService.archive_with_check(id, actor_id=get_actor_id())
        flash("Firma başarıyla arşive kaldırıldı.", 'success')
    except ValidationError as e:
        flash(str(e), 'warning')
    except Exception as e:
        current_app.logger.error(f"Arşivleme hatası (ID: {id}): {str(e)}")
        flash("İşlem sırasında sistemsel bir hata oluştu.", 'danger')
    return redirect(url_for('firmalar.index'))

# -------------------------------------------------------------------------
# 8. Firmayı Aktifleştir (Arşivden Geri Al)
# -------------------------------------------------------------------------
@firmalar_bp.route('/aktiflestir/<int:id>', methods=['POST'])
@login_required
def aktiflestir(id):
    try:
        FirmaService.update(id, {'is_active': True}, actor_id=get_actor_id())
        flash("Firma başarıyla tekrar aktif hale getirildi.", "success")
    except Exception as e:
        current_app.logger.error(f"Aktifleştirme hatası (ID: {id}): {str(e)}")
        flash("İşlem başarısız oldu.", "danger")
    return redirect(url_for('firmalar.pasif_index'))

# -------------------------------------------------------------------------
# 9. İmza Yetkisi Kontrolü
# -------------------------------------------------------------------------
@firmalar_bp.route('/imza-kontrol/<int:id>', methods=['POST'])
@login_required
def imza_kontrol(id):
    try:
        FirmaService.update(
            id, 
            {'imza_yetkisi_kontrol_edildi': True, 'imza_yetkisi_kontrol_tarihi': date.today()}, 
            actor_id=get_actor_id()
        )
        flash("İmza yetkisi başarıyla onaylandı.", "success")
    except ValidationError as e:
        flash(str(e), "danger")
    except Exception as e:
        current_app.logger.error(f"İmza kontrol hatası (ID: {id}): {str(e)}")
        flash("Onay işlemi yapılamadı.", "danger")
    return redirect(url_for('firmalar.bilgi', id=id))


# -------------------------------------------------------------------------
# 10. Excel Dışa Aktar / İçe Yükle
# -------------------------------------------------------------------------
@firmalar_bp.route('/excel-disari-aktar', methods=['GET'])
@login_required
def excel_disari_aktar():
    firmalar = Firma.query.filter_by(is_active=True).order_by(Firma.firma_adi).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Firmalar'

    headers = [
        'Firma Adi', 'Yetkili Adi', 'Telefon', 'Eposta', 'Iletisim Bilgileri',
        'Vergi Dairesi', 'Vergi No', 'Musteri Mi', 'Tedarikci Mi', 'Aktif Mi'
    ]
    ws.append(headers)

    for f in firmalar:
        ws.append([
            f.firma_adi,
            f.yetkili_adi,
            f.telefon,
            f.eposta,
            f.iletisim_bilgileri,
            f.vergi_dairesi,
            f.vergi_no,
            'Evet' if f.is_musteri else 'Hayir',
            'Evet' if f.is_tedarikci else 'Hayir',
            'Evet' if f.is_active else 'Hayir',
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name=f"firmalar_{date.today().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@firmalar_bp.route('/excel-ice-yukle', methods=['POST'])
@login_required
def excel_ice_yukle():
    file = request.files.get('excel_file')
    if not file or not file.filename:
        flash('Lütfen bir Excel dosyası seçiniz.', 'warning')
        return redirect(url_for('firmalar.index'))

    if not file.filename.lower().endswith('.xlsx'):
        flash('Sadece .xlsx uzantılı dosyalar destekleniyor.', 'danger')
        return redirect(url_for('firmalar.index'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as exc:
        flash(f'Excel dosyası okunamadı: {exc}', 'danger')
        return redirect(url_for('firmalar.index'))

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            firma_adi = (row[0] or '').strip() if row[0] else ''
            yetkili_adi = (row[1] or '').strip() if row[1] else ''
            telefon = (row[2] or '').strip() if row[2] else None
            eposta = (row[3] or '').strip() if row[3] else None
            iletisim_bilgileri = (row[4] or '').strip() if row[4] else ''
            vergi_dairesi = (row[5] or '').strip() if row[5] else ''
            vergi_no = (row[6] or '').strip() if row[6] else ''
            is_musteri = _to_bool(row[7], default=True)
            is_tedarikci = _to_bool(row[8], default=False)
            is_active = _to_bool(row[9], default=True)

            if not vergi_no:
                skipped += 1
                continue

            firma_adi = firma_adi or f'Firma {vergi_no}'
            yetkili_adi = yetkili_adi or 'Belirtilmedi'
            iletisim_bilgileri = iletisim_bilgileri or (telefon or '-')
            vergi_dairesi = vergi_dairesi or 'Belirtilmedi'

            mevcut = Firma.query.filter_by(vergi_no=vergi_no).first()
            if mevcut:
                mevcut.firma_adi = firma_adi
                mevcut.yetkili_adi = yetkili_adi
                mevcut.telefon = telefon
                mevcut.eposta = eposta
                mevcut.iletisim_bilgileri = iletisim_bilgileri
                mevcut.vergi_dairesi = vergi_dairesi
                mevcut.is_musteri = is_musteri
                mevcut.is_tedarikci = is_tedarikci
                mevcut.is_active = is_active
                updated += 1
            else:
                yeni = Firma(
                    firma_adi=firma_adi,
                    yetkili_adi=yetkili_adi,
                    telefon=telefon,
                    eposta=eposta,
                    iletisim_bilgileri=iletisim_bilgileri,
                    vergi_dairesi=vergi_dairesi,
                    vergi_no=vergi_no,
                    is_musteri=is_musteri,
                    is_tedarikci=is_tedarikci,
                    is_active=is_active,
                    bakiye=Decimal('0'),
                )
                db.session.add(yeni)
                created += 1
        except Exception as exc:
            errors.append(f'Satır {row_idx}: {exc}')

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f'Excel içe aktarma başarısız: {exc}', 'danger')
        return redirect(url_for('firmalar.index'))

    flash(
        f'Excel içe aktarım tamamlandı. Yeni: {created}, Güncellenen: {updated}, Atlanan: {skipped}.',
        'success'
    )
    if errors:
        flash('Bazı satırlar işlenemedi: ' + ' | '.join(errors[:5]), 'warning')

    return redirect(url_for('firmalar.index'))