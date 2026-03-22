# app/db_menu/routes.py
from flask import request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from datetime import datetime
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect as sa_inspect, MetaData, Table, select

from app import db
from app.db_menu import db_menu_bp
from app.utils import admin_required


# ---------------------------------------------------------------------------
# Yardımcı fonksiyonlar
# ---------------------------------------------------------------------------

def _excel_safe_sheet_name(name, index):
    cleaned = ''.join(ch for ch in str(name) if ch not in '[]:*?/\\')
    cleaned = cleaned[:31].strip() or f'Sheet{index}'
    return cleaned


def _normalize_excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    if isinstance(value, bytes):
        try:
            return value.decode('utf-8')
        except Exception:
            return str(value)
    return value


def _to_table_row_dicts(sheet, valid_columns):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []

    headers = [str(col).strip() if col is not None else '' for col in header_row]
    usable_headers = [h for h in headers if h in valid_columns]
    if not usable_headers:
        return []

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        data = {}
        has_value = False
        for idx, header in enumerate(headers):
            if header not in usable_headers:
                continue
            value = row[idx] if idx < len(row) else None
            data[header] = value
            if value not in (None, ''):
                has_value = True
        if has_value:
            rows.append(data)
    return rows


# ---------------------------------------------------------------------------
# Context processor — tüm şablonlara tablo listesi enjekte eder
# ---------------------------------------------------------------------------

@db_menu_bp.app_context_processor
def inject_admin_db_backup_tables():
    if not current_user.is_authenticated:
        return {}
    try:
        if not current_user.is_admin():
            return {}
    except Exception:
        return {}

    try:
        table_names = sorted(
            t for t in sa_inspect(db.engine).get_table_names()
            if t != 'alembic_version'
        )
    except Exception:
        table_names = []

    return {'db_backup_tables': table_names}


# ---------------------------------------------------------------------------
# Rotalar
# ---------------------------------------------------------------------------

@db_menu_bp.route('/yedek-excel', methods=['POST'])
@login_required
@admin_required
def db_yedek_excel():
    secilen_tablolar = request.form.getlist('tables')
    if not secilen_tablolar:
        flash('Lütfen Excel yedeği için en az bir tablo seçin.', 'warning')
        return redirect(request.referrer or url_for('main.index'))

    try:
        mevcut_tablolar = set(sa_inspect(db.engine).get_table_names())
    except Exception:
        flash('Tablo listesi alınamadı. Lütfen tekrar deneyin.', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    tablolar = [
        t for t in secilen_tablolar
        if t in mevcut_tablolar and t != 'alembic_version'
    ]

    if not tablolar:
        flash('Geçerli tablo seçimi bulunamadı.', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    bilgi = wb.create_sheet('Yedek_Bilgi')
    bilgi.append(['Olusturma Zamanı', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    bilgi.append(['Olusturan Kullanıcı', current_user.username])
    bilgi.append(['Tablo Sayısı', len(tablolar)])
    bilgi.append(['Tablolar', ', '.join(tablolar)])

    metadata = MetaData()

    try:
        for idx, tablo_adi in enumerate(tablolar, start=1):
            tablo = Table(tablo_adi, metadata, autoload_with=db.engine)
            sayfa = wb.create_sheet(_excel_safe_sheet_name(tablo_adi, idx))

            sutunlar = [col.name for col in tablo.columns]
            sayfa.append(sutunlar)

            satirlar = db.session.execute(select(tablo)).mappings().all()
            for satir in satirlar:
                sayfa.append([_normalize_excel_value(satir.get(sutun)) for sutun in sutunlar])
    except Exception as exc:
        flash(f'Excel yedeği alınırken hata oluştu: {exc}', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    dosya_adi = f"db_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=dosya_adi,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@db_menu_bp.route('/yedek-excel-yukle', methods=['POST'])
@login_required
@admin_required
def db_yedek_excel_yukle():
    secilen_tablolar = request.form.getlist('restore_tables')
    file = request.files.get('backup_file')

    if not file or not file.filename:
        flash('Lütfen geri yüklemek için bir Excel dosyası seçin.', 'warning')
        return redirect(request.referrer or url_for('main.index'))

    if not file.filename.lower().endswith('.xlsx'):
        flash('Sadece .xlsx uzantılı dosyalar destekleniyor.', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    if not secilen_tablolar:
        flash('Geri yükleme için en az bir tablo seçmelisiniz.', 'warning')
        return redirect(request.referrer or url_for('main.index'))

    try:
        mevcut_tablolar = set(sa_inspect(db.engine).get_table_names())
    except Exception:
        flash('Tablo listesi alınamadı. Lütfen tekrar deneyin.', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    tablolar = [
        t for t in secilen_tablolar
        if t in mevcut_tablolar and t != 'alembic_version'
    ]
    if not tablolar:
        flash('Geçerli tablo seçimi bulunamadı.', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    try:
        workbook = load_workbook(file, data_only=True)
    except Exception as exc:
        flash(f'Excel dosyası okunamadı: {exc}', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    metadata = MetaData()
    toplam_silinen = 0
    toplam_eklenen = 0
    islenen_tablo = 0

    try:
        for idx, tablo_adi in enumerate(tablolar, start=1):
            if tablo_adi in workbook.sheetnames:
                sayfa = workbook[tablo_adi]
            else:
                safe_name = _excel_safe_sheet_name(tablo_adi, idx)
                if safe_name in workbook.sheetnames:
                    sayfa = workbook[safe_name]
                else:
                    continue

            tablo = Table(tablo_adi, metadata, autoload_with=db.engine)
            sutunlar = {col.name for col in tablo.columns}
            satirlar = _to_table_row_dicts(sayfa, sutunlar)

            silme_sonucu = db.session.execute(tablo.delete())
            toplam_silinen += silme_sonucu.rowcount or 0

            if satirlar:
                db.session.execute(tablo.insert(), satirlar)
                toplam_eklenen += len(satirlar)

            islenen_tablo += 1

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f'Excel geri yükleme başarısız: {exc}', 'danger')
        return redirect(request.referrer or url_for('main.index'))

    flash(
        f'Geri yükleme tamamlandı. Tablo: {islenen_tablo}, Silinen kayıt: {toplam_silinen}, Eklenen kayıt: {toplam_eklenen}.',
        'success'
    )
    return redirect(request.referrer or url_for('main.index'))
